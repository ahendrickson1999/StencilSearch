from openpyxl import load_workbook
import json

wb = load_workbook("StencilData.xlsx")
ws = wb.active

headers = [cell.value for cell in ws[1]]

data = []

for row in ws.iter_rows(min_row=2, values_only=True):
    record = {}

    for header, value in zip(headers, row):
        if value is None:
            value = ""

        record[header] = value

    data.append(record)

with open("stencils.json", "w") as f:
    json.dump(data, f, indent=4)

print("Finished!")